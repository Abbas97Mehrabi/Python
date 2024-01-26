from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = load_workbook('PythonTutorial/Python-Excel/Grades.xlsx')
ws = wb.active
print(ws['A2'].value)
#ws['A2'].value = "Emin"
#wb.create_sheet("Kadir")
#print(wb.sheetnames)

#for row in range(1, 7):
 #   for col in range(1, 5):
  #      char = get_column_letter(col)
   #     print(ws[char + str(row)].value)

#ws.merge_cells("A1:D2")

#ws.insert_rows(7)
#ws.delete_rows(7)
#ws.insert_cols(2)
#ws.delete_cols(2)

#ws.move_range("C1:D11", rows=2, cols=2)

for col in range(1, 6):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

wb.save('PythonTutorial/Python-Excel/Grades.xlsx')

#openpyxl.readthedocs.io is site of this  documentation